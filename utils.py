import pandas as pd
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import os

def find_best_sheet(file_path):
    """
    Najde list s největší datovou plochou (počet řádků * počet sloupců).
    Sdílená funkce pro autodetekci listu.
    """
    try:
        xl = pd.ExcelFile(file_path)
    except Exception as e:
        raise ValueError(f"Nelze otevřít Excel soubor: {e}")

    best_sheet = None
    max_area = -1
    
    for sheet_name in xl.sheet_names:
        # Pandas read_excel je nákladné, pro optimalizaci by šlo použít openpyxl read_only,
        # ale pro běžné velikosti souborů je toto OK a robustní.
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        area = df.shape[0] * df.shape[1]
        
        if area > max_area and not df.empty:
            max_area = area
            best_sheet = sheet_name
            
    if best_sheet is None:
        raise ValueError("Nenašel jsem žádný list s daty.")
        
    print(f"   -> Vybrán list: '{best_sheet}' (Plocha: {max_area} buněk)")
    return best_sheet

def create_excel_table(file_path, sheet_name, table_name):
    """
    Otevře existující Excel, na daném listu vytvoří formátovanou tabulku (ListObject)
    přes celou oblast dat a soubor uloží.
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            print(f"⚠️ Varování: List '{sheet_name}' v souboru neexistuje, tabulka nevytvořena.")
            return

        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column

        # Pokud je list prázdný nebo má jen hlavičku
        if max_row < 2:
            print(f"⚠️ List '{sheet_name}' má málo dat, tabulka nevytvořena.")
            return

        # Získáme písmeno posledního sloupce (např. 4 -> D)
        last_col_letter = get_column_letter(max_col)
        
        # Definice rozsahu např. "A1:D150"
        ref = f"A1:{last_col_letter}{max_row}"
        
        # Vytvoření tabulky
        tab = Table(displayName=table_name, ref=ref)
        
        # Styl (modrý pruhovaný - standard)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        
        # Pokud už tabulka se stejným názvem existuje (při přepsání), openpyxl by mohl spadnout
        # Zde jednoduše přidáme novou. Excel si případně poradí (nebo starou přepíšeme v Python logic předtím)
        ws.add_table(tab)
        
        wb.save(file_path)
        print(f"   -> Tabulka '{table_name}' vytvořena a uložena.")
        
    except Exception as e:
        print(f"❌ Chyba při formátování tabulky v {file_path}: {e}")