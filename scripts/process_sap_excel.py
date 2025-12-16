import pandas as pd
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import os
import argparse

INPUT_DIR = "sklady_porovnani/input"
OUTPUT_DIR = "sklady_porovnani/output"
OUTPUT_FILENAME = "SAP.xlsx"

REQUIRED_COLUMNS = ["Material", "Material description", "Batch", "Total Quantity", "Storage location"]
OPTIONAL_COLUMNS = ["Plant"]
STORAGE_LOCATIONS = {"F010", "F070"}
TABLE_NAME = "tbl_SAP"
SHEET_NAME = "SAP"

def find_best_sheet(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    best_sheet = None
    max_area = 0
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        if ws.max_row < 1:
            continue
        if all(cell.value is None for cell in ws[1]):
            continue
        area = ws.max_row * ws.max_column
        if area > max_area:
            best_sheet = sheet
            max_area = area
    if not best_sheet:
        raise ValueError("Nenašel jsem list s hlavičkou v 1. řádku.")
    return best_sheet

def standardize_columns(df):
    col_map = {col.lower(): col for col in df.columns}
    missing = [col for col in REQUIRED_COLUMNS if col.lower() not in col_map]
    if missing:
        raise ValueError(f"Chybí sloupce: {', '.join(missing)}")
    return col_map

def process_dataframe(df, col_map):
    df.columns = [col.strip() for col in df.columns]

    df = df[df[col_map["storage location"].lower()].isin(STORAGE_LOCATIONS)]

    if "plant" in col_map:
        df = df.drop(columns=[col_map["plant"]])
    df = df.drop(columns=[col_map["storage location"]])

    ordered = [
        col_map["material"],
        col_map["material description"],
        col_map["batch"],
        col_map["total quantity"]
    ]
    remaining = [col for col in df.columns if col not in ordered]
    df = df[ordered + remaining]
    df = df.iloc[:, :4]

    df.columns = ["Material", "Název", "Batch", "Mnozstvi_SAP"]
    df["Mnozstvi_SAP"] = pd.to_numeric(df["Mnozstvi_SAP"], errors="coerce").fillna(0)
    df = df.sort_values(by="Mnozstvi_SAP", ascending=False, kind="mergesort")

    if df.shape[0] > 1:
        df = df.drop(df.index[0])

    return df

def write_to_excel(df, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    for i, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=i, value=col)
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    last_row = len(df) + 1
    table_range = f"A1:D{last_row}"
    table = Table(displayName=TABLE_NAME, ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style

    if ws._tables:
        ws._tables.clear()

    ws.add_table(table)

    if SHEET_NAME in wb.sheetnames and wb.sheetnames[0] != SHEET_NAME:
        del wb[SHEET_NAME]

    wb.save(output_path)

def process_file(filepath, output_dir):
    best_sheet = find_best_sheet(filepath)
    df = pd.read_excel(filepath, sheet_name=best_sheet, dtype=str)
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    col_map = standardize_columns(df)
    df_processed = process_dataframe(df, col_map)
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, OUTPUT_FILENAME)
    write_to_excel(df_processed, output_path)

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["all", "single"], required=True)
    parser.add_argument("--file_path", help="Path to the Excel file for single mode")
    args = parser.parse_args()

    if args.mode == "all":
        files = list(Path(INPUT_DIR).glob("*.xlsx"))
        if not files:
            print("Žádné .xlsx soubory ve složce input.")
            return
        for file in files:
            try:
                process_file(file, OUTPUT_DIR)
            except Exception as e:
                print(f"Chyba při zpracování {file}: {e}")
    elif args.mode == "single":
        if not args.file_path:
            print("Pro režim 'single' je nutné zadat --file_path")
            return
        file = Path(args.file_path)
        if not file.exists():
            print(f"Soubor {args.file_path} neexistuje.")
            return
        try:
            process_file(file, OUTPUT_DIR)
        except Exception as e:
            print(f"Chyba při zpracování {file}: {e}")

if __name__ == "__main__":
    main()
