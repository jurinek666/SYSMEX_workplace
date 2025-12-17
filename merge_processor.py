import pandas as pd
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import sys
import os
import hashlib

def calculate_hash(row):
    """
    Vytvoří SHA-256 hash z Material a Batch.
    Pravidla: UPPER(TRIM(Material)) + '|' + UPPER(TRIM(Batch))
    """
    # Získání hodnot, ošetření NaN -> prázdný string
    mat = str(row['Material']).strip().upper() if pd.notna(row['Material']) else ""
    batch = str(row['Batch']).strip().upper() if pd.notna(row['Batch']) else ""
    
    # Sestavení řetězce
    raw_string = f"{mat}|{batch}"
    
    # Hashování
    return hashlib.sha256(raw_string.encode('utf-8')).hexdigest()

def create_excel_table(ws, sheet_name, table_name):
    """Pomocná funkce pro vytvoření Excel tabulky"""
    max_row = ws.max_row
    max_col = ws.max_column
    # Převod čísla sloupce na písmeno (např. 5 -> E)
    from openpyxl.utils import get_column_letter
    col_letter = get_column_letter(max_col)
    
    ref = f"A1:{col_letter}{max_row}"
    
    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

def process_merge():
    print("--- Začínám slučování a finální úpravy ---")
    
    # Cesty
    sap_path = "sklady_porovnani/output/SAP.xlsx"
    raben_path = "sklady_porovnani/output/RABEN.xlsx"
    output_path = "sklady_porovnani/input/POROVNANI_SKLADU.xlsx"
    
    # Kontrola vstupů
    if not os.path.exists(sap_path) or not os.path.exists(raben_path):
        print("❌ Chybí vstupní soubory v output složce (SAP.xlsx nebo RABEN.xlsx).")
        sys.exit(1)

    try:
        # 1. Načtení dat
        df_sap = pd.read_excel(sap_path, sheet_name="SAP")
        df_raben = pd.read_excel(raben_path, sheet_name="RABEN")
        
        print(f"Načteno: SAP ({len(df_sap)} řádků), RABEN ({len(df_raben)} řádků)")

        # --- LOGIKA ÚPRAV PRO RABEN ---
        
        # A. Filtrace obalů (Maska P*** nebo P****)
        # Regex: Začátek(^) P, následují 3 nebo 4 číslice(\d{3,4}), Konec($)
        initial_count = len(df_raben)
        mask_packaging = df_raben['Material'].astype(str).str.match(r'^P\d{3,4}$', case=False)
        df_raben = df_raben[~mask_packaging].copy()
        print(f"Filtr obalů (P...): Odstraněno {initial_count - len(df_raben)} řádků.")

        # B. Přepočet množství pro ZE001906 (* 50)
        mask_item = df_raben['Material'] == 'ZE001906'
        count_items = mask_item.sum()
        if count_items > 0:
            df_raben.loc[mask_item, 'Mnozstvi_RABEN'] = df_raben.loc[mask_item, 'Mnozstvi_RABEN'] * 50
            print(f"Přepočet: Upraveno množství u {count_items} řádků (ZE001906).")

        # --- LOGIKA HASH (PRO OBĚ TABULKY) ---
        print("Generuji HASH sloupce...")
        df_sap['HASH'] = df_sap.apply(calculate_hash, axis=1)
        df_raben['HASH'] = df_raben.apply(calculate_hash, axis=1)

        # --- ZÁPIS DO POROVNANI_SKLADU.xlsx ---
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_sap.to_excel(writer, sheet_name='SAP', index=False)
            df_raben.to_excel(writer, sheet_name='RABEN', index=False)
            
        # --- FORMÁTOVÁNÍ TABULEK ---
        wb = openpyxl.load_workbook(output_path)
        
        # Formátování listu SAP
        create_excel_table(wb["SAP"], "SAP", "tbl_SAP")
        
        # Formátování listu RABEN
        create_excel_table(wb["RABEN"], "RABEN", "tbl_RABEN")
        
        wb.save(output_path)
        print(f"✅ HOTOVO. Master soubor vytvořen: {output_path}")

    except Exception as e:
        print(f"❌ Chyba při slučování: {e}")
        sys.exit(1)

if __name__ == "__main__":
    process_merge()