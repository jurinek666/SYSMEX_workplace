import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os
import sys

# --- KONFIGURACE ---
INPUT_FILE = "sklady_porovnani/input/POROVNANI_SKLADU.xlsx"
OUTPUT_FILE = "sklady_porovnani/output/vysledky.xlsx"

# Barvy pro podmíněné formátování (HEX kód bez #)
COLOR_OK = "C6EFCE"       # Světle zelená
COLOR_MANKO = "FFC7CE"    # Světle červená
COLOR_PREBYTEK = "FFEB9C" # Oranžová

def create_hash_col(df):
    """
    Vytvoří pomocný sloupec HASH spojením Material a Batch.
    """
    mat = df['Material'].astype(str).str.strip().replace('nan', '')
    bat = df['Batch'].astype(str).str.strip().replace('nan', '')
    return mat + "_" + bat

def compare_data():
    print(f"--- Spouštím porovnání dat: {INPUT_FILE} ---")

    if not os.path.exists(INPUT_FILE):
        print(f"❌ CHYBA: Vstupní soubor neexistuje: {INPUT_FILE}")
        # V rámci workflow to může znamenat, že předchozí krok selhal
        sys.exit(1)

    try:
        # 1. Načtení dat (vše jako text, abychom neztratili nuly)
        print("Načítám SAP a RABEN...")
        df_sap = pd.read_excel(INPUT_FILE, sheet_name="SAP", dtype=str)
        df_raben = pd.read_excel(INPUT_FILE, sheet_name="RABEN", dtype=str)

        # 2. Příprava klíčů (HASH)
        df_sap['HASH'] = create_hash_col(df_sap)
        df_raben['HASH'] = create_hash_col(df_raben)

        # Převod množství na čísla pro výpočet
        df_sap['Mnozstvi_SAP'] = pd.to_numeric(df_sap['Mnozstvi_SAP'], errors='coerce').fillna(0)
        df_raben['Mnozstvi_RABEN'] = pd.to_numeric(df_raben['Mnozstvi_RABEN'], errors='coerce').fillna(0)

        # 3. Spojení (Full Outer Join)
        print("Provádím párování (Outer Join)...")
        df_merged = pd.merge(
            df_sap, 
            df_raben, 
            on='HASH', 
            how='outer', 
            suffixes=('_SAP', '_RABEN')
        )

        # 4. Konsolidace sloupců (Coalesce - když chybí v SAP, vezmi z RABEN a naopak)
        print("Konsoliduji data...")
        df_merged['Material'] = df_merged['Material_SAP'].combine_first(df_merged['Material_RABEN'])
        df_merged['Nazev'] = df_merged['Nazev_SAP'].combine_first(df_merged['Nazev_RABEN'])
        df_merged['Batch'] = df_merged['Batch_SAP'].combine_first(df_merged['Batch_RABEN'])

        # Doplnění 0 tam, kde data chybí (např. zboží je jen v SAPu -> RABEN = 0)
        df_merged['Mnozstvi_SAP'] = df_merged['Mnozstvi_SAP'].fillna(0)
        df_merged['Mnozstvi_RABEN'] = df_merged['Mnozstvi_RABEN'].fillna(0)

        # 5. Výpočty a STAV
        df_merged['Rozdil'] = df_merged['Mnozstvi_RABEN'] - df_merged['Mnozstvi_SAP']

        def urcit_stav(row):
            diff = row['Rozdil']
            if diff == 0:
                return "STAV OK"
            elif diff > 0:
                return "RABEN manko"    # Dle zadání: Rozdíl > 0 (RABEN má víc? Pozor, ověř logiku "manko/přebytek")
                # Obvykle: Manko = chybí fyzicky (RABEN < SAP). 
                # Zde dle tvého zadání: Rozdil = RABEN - SAP.
                # Pokud Rozdil > 0 (RABEN > SAP), je to "RABEN manko" (dle tvého promptu).
                # (Technicky je to přebytek ve skladu, ale držím se přesně tvého zadání textů).
            else:
                return "RABEN přebytek" # Dle zadání: Rozdil < 0

        df_merged['STAV'] = df_merged.apply(urcit_stav, axis=1)

        # 6. Finální výběr
        final_cols = ["Material", "Nazev", "Batch", "Mnozstvi_SAP", "Mnozstvi_RABEN", "Rozdil", "STAV"]
        df_final = df_merged[final_cols]

        # 7. Uložení
        print(f"Ukládám do: {OUTPUT_FILE}")
        os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
        
        with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
            df_final.to_excel(writer, sheet_name='Vysledky', index=False)

        # 8. Formátování
        apply_formatting(OUTPUT_FILE)
        
        print("✅ Hotovo. Report vygenerován.")

    except Exception as e:
        print(f"❌ Chyba při porovnávání: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

def apply_formatting(file_path):
    """
    Otevře Excel a aplikuje podmíněné formátování řádků.
    """
    print("Aplikuji barevné formátování...")
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Vysledky']

    # Styly
    fill_ok = PatternFill(start_color=COLOR_OK, end_color=COLOR_OK, fill_type="solid")
    fill_manko = PatternFill(start_color=COLOR_MANKO, end_color=COLOR_MANKO, fill_type="solid")
    fill_prebytek = PatternFill(start_color=COLOR_PREBYTEK, end_color=COLOR_PREBYTEK, fill_type="solid")
    
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    # Barvení řádků podle sloupce STAV (7. sloupec, index G)
    for row in ws.iter_rows(min_row=2):
        stav_cell = row[6] # Index 6 = 7. sloupec (G)
        stav_val = stav_cell.value 
        
        target_fill = None
        if stav_val == "STAV OK":
            target_fill = fill_ok
        elif stav_val == "RABEN manko":
            target_fill = fill_manko
        elif stav_val == "RABEN přebytek":
            target_fill = fill_prebytek
            
        if target_fill:
            for cell in row:
                cell.fill = target_fill

    # Autofit šířky sloupců
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(file_path)

if __name__ == "__main__":
    compare_data()