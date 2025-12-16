# path: sklady_porovnani/scripts/transform_sap.py
from __future__ import annotations
import argparse
from pathlib import Path
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo

REQUIRED_IN = ["Material", "Material description", "Batch", "Total Quantity", "Storage location"]
FINAL_ORDER = ["Material", "Material description", "Batch", "Total Quantity"]
RENAME_MAP = {"Material description": "Název", "Total Quantity": "Mnozstvi_SAP"}
ALLOWED_STORAGE = {"F010", "F070"}  # záměrně case-sensitive

def _ci_map(cols: list[str]) -> dict[str, str]:
    return {str(c).strip().lower(): str(c).strip() for c in cols}

def _require(df: pd.DataFrame, cols: list[str]):
    ci = _ci_map(df.columns.tolist())
    missing = [c for c in cols if c.lower() not in ci]
    if missing:
        raise SystemExit(f"Chybí sloupce: {', '.join(missing)}")

def _resolve(df: pd.DataFrame, name: str) -> str:
    ci = _ci_map(df.columns.tolist())
    key = name.strip().lower()
    if key not in ci:
        raise SystemExit(f"Sloupec '{name}' nenalezen")
    return ci[key]

def read_best_sheet(path: Path) -> pd.DataFrame:
    xl = pd.ExcelFile(path, engine="openpyxl")
    best_name, best_score = None, -1
    for name in xl.sheet_names:
        df = xl.parse(name, header=0)
        if df.empty: continue
        non_empty_headers = sum(1 for c in df.columns if str(c).strip() != "")
        if non_empty_headers == 0: continue
        score = df.shape[0] * df.shape[1]
        if score > best_score:
            best_score, best_name = score, name
    if best_name is None:
        raise SystemExit("Nenašel jsem list s hlavičkou v řádku 1 a daty.")
    return xl.parse(best_name, header=0)

def transform(df: pd.DataFrame) -> pd.DataFrame:
    _require(df, REQUIRED_IN)
    col_storage = _resolve(df, "Storage location")
    col_plant = _resolve(df, "Plant") if "plant" in _ci_map(df.columns) else None

    df = df.copy()
    df[col_storage] = df[col_storage].astype(str).str.strip()
    df = df[df[col_storage].isin(ALLOWED_STORAGE)]

    drop_cols = [col_storage] + ([col_plant] if col_plant else [])
    df = df.drop(columns=drop_cols, errors="ignore")

    resolved_order = [_resolve(df, c) for c in FINAL_ORDER]
    rest = [c for c in df.columns if c not in resolved_order]
    df = df[resolved_order + rest]

    rename_resolved = {_resolve(df, k): v for k, v in RENAME_MAP.items()}
    df = df.rename(columns=rename_resolved)

    qty_col = "Mnozstvi_SAP"
    df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
    df = df.sort_values(by=[qty_col], ascending=False, kind="mergesort")

    if len(df) > 0:
        df = df.iloc[1:, :]  # smazat první datový řádek po sortu

    df = df.iloc[:, :4].copy()
    return df.reset_index(drop=True)

def write_excel(df: pd.DataFrame, out_path: Path, sheet_name: str = "SAP"):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    wb = load_workbook(out_path)
    ws: Worksheet = wb[sheet_name]
    last_row = ws.max_row
    ref = f"A1:D{last_row}"
    # odeber duplicitní tabulku stejného jména
    ws._tables = [t for t in ws._tables if t.displayName.lower() != "tbl_sap"]
    tbl = Table(displayName="tbl_SAP", ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)
    wb.save(out_path)

def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    args = ap.parse_args(argv)

    src, dst = Path(args.input), Path(args.output)
    if not src.exists():
        raise SystemExit(f"Soubor neexistuje: {src}")

    df = read_best_sheet(src)
    df_out = transform(df)
    write_excel(df_out, dst)
    print(f"[OK] -> {dst}")
    return 0

if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))